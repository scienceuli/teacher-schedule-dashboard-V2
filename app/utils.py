import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from functools import wraps
from flask import session, redirect, url_for, flash

def is_valid_teacher(name):
    return isinstance(name, str) and name.count(",") == 1 and all(part.strip() for part in name.split(","))


def get_file(data_dir, extensions):
    
    data_files = [os.path.join(data_dir, filename) for filename in os.listdir(data_dir) if allowed_file(filename, extensions)]

    if not data_files:
        return None  # No file found

    print('data_files:', data_files)

    data_file = max(data_files, key=os.path.getmtime)

    print('newest data_file:', data_file)

    return data_file

def allowed_file(filename, extensions):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in extensions and not filename.startswith("~$")


def create_folder(folder_path):
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

def rename_columns(x):
    header = x[0]
    subheader = x[1]
    if header.startswith("Sonderaufgaben"):
        return "Sonderaufgaben"
    if header.startswith("Ags") and subheader == "AG":
        return "Ags-AG"
    if header.startswith("Ags") and subheader == "Std":
        return "Ags-Std"
    if header.startswith("Poolstd") and subheader == "Bg":
        return "Poolstd-Bg"
    if header.startswith("Poolstd") and subheader == "Std":
        return "Poolstd-Std"
    return header

def style_excel_output(wb, ws_name, columns, highlight_column=None, highlight_cell=None):
    ws = wb[ws_name]

    # Freeze header row
    ws.freeze_panes = "A2"

    # Style header
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get column index (number)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Optional: Highlight Delta > 0
    if highlight_column:
        
        for row in ws.iter_rows(min_row=2, min_col=columns.index(highlight_column) + 1, max_col=columns.index(highlight_column) + 1):
            for cell in row:
                try:
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                except:
                    continue

    if highlight_cell:
        target_row_label = highlight_cell.get('row')
        target_col_label = highlight_cell.get('column')
        target_row = None
        target_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == target_col_label:
                target_col = col_idx
                break
        
        for row_idx in range(2, ws.max_row + 1):  # assuming headers in row 1
            if ws.cell(row=row_idx, column=1).value == target_row_label:
                target_row = row_idx
                break
        if target_row and target_col:
            cell = ws.cell(row=target_row, column=target_col)
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid") 
            elif isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            

    return wb

def set_alternating_column_background(ws, start_row=1, start_col=1, step=3, color1="FFFFCC", color2="FFFFFF"):
    
    max_col = ws.max_column
    max_row = ws.max_row

    for group_start in range(start_col, max_col + 1, step):
        group_end = min(group_start + step - 1, max_col)
        # Alternate color block
        color = color1 if ((group_start - start_col) // step) % 2 == 0 else color2
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        border = Border(right=Side(style='thin'),)

        for col in range(group_start, group_end + 1):
            for row in range(start_row, max_row + 1):
                ws.cell(row=row, column=col).fill = fill
        
        for row in range(start_row, max_row + 1):
            ws.cell(row=row, column=group_start+step-1).border = border
            # ws.cell(row=row, column=group_end).border = border

    return ws

def insert_excel_rows(ws, string, number_of_rows=1):
    header_font = Font(bold=True)
    for _ in range(number_of_rows):
        ws.insert_rows(1)
        
    ws.cell(row=1, column=1).value = string
    ws.cell(row=1, column=1).font = header_font

def set_size(ws, orientation="landscape"):
    if orientation == "landscape":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1



def login_required(view_func):
    @wraps(view_func)
    def wrapped_view(*args, **kwargs):
        if 'username' not in session:
            flash('Please log in first.', 'warning')
            return redirect(url_for('accounts.login'))
        return view_func(*args, **kwargs)
    return wrapped_view


def convert_empty_string_to_zero(value):
    if type(value) == str and value == "":
        return 0
    elif type(value) == str:
        return float(value)
        
    return value








